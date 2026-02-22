# -*- coding: utf-8 -*-
"""
Mistral API 백업 파서 테스트
- 교정성적서 Excel 데이터를 LLM에 넘겨서 구조화 추출
- 규칙 기반 파싱 vs LLM 파싱 결과 비교
"""

import sys
import io
import json
import time
import requests
import openpyxl
from io import BytesIO

if sys.platform == 'win32' and hasattr(sys.stdout, 'buffer') and not isinstance(sys.stdout, io.TextIOWrapper):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from ktools_login import ktools_login
from ktools_수집 import fetch_all
from ktools_성적서 import make_api_acpt_no, download_cert_excel, parse_cert_excel

# =============================================================================
# Mistral API 설정
# =============================================================================
MISTRAL_API_KEY = '8GmClOHJYZr3PpZUAunyzQykOJPghT8D'
MISTRAL_URL = 'https://api.mistral.ai/v1/chat/completions'
MISTRAL_MODEL = 'mistral-small-latest'  # 비용 효율 + 빠른 속도

# k-tools 설정
USER_ID = 'hicor'
USER_PWD = 'dlacodnr1!'
BASE_URL = 'https://k-tools.ktl.re.kr'


def excel_to_text(excel_bytes):
    """Excel 바이트 → 시트별 텍스트 변환 (LLM 입력용)"""
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    result = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        lines = [f'=== Sheet: {sn} ===']
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() if v is not None else '' for v in row]
            # 빈 행 스킵
            if not any(vals):
                continue
            lines.append(' | '.join(vals))
        result.append('\n'.join(lines))

    wb.close()
    return '\n\n'.join(result)


def call_mistral(prompt, system_prompt=None, temperature=0.0, max_tokens=2000):
    """Mistral API 호출 → 응답 텍스트 반환"""
    messages = []
    if system_prompt:
        messages.append({'role': 'system', 'content': system_prompt})
    messages.append({'role': 'user', 'content': prompt})

    payload = {
        'model': MISTRAL_MODEL,
        'messages': messages,
        'temperature': temperature,
        'max_tokens': max_tokens,
        'response_format': {'type': 'json_object'},
    }

    headers = {
        'Authorization': f'Bearer {MISTRAL_API_KEY}',
        'Content-Type': 'application/json',
    }

    t0 = time.time()
    res = requests.post(MISTRAL_URL, json=payload, headers=headers, timeout=60)
    elapsed = time.time() - t0

    if res.status_code != 200:
        raise Exception(f'Mistral API 에러 {res.status_code}: {res.text[:200]}')

    data = res.json()
    content = data['choices'][0]['message']['content']
    usage = data.get('usage', {})

    return {
        'content': content,
        'elapsed': elapsed,
        'input_tokens': usage.get('prompt_tokens', 0),
        'output_tokens': usage.get('completion_tokens', 0),
    }


SYSTEM_PROMPT = """You are a calibration certificate data extraction assistant.
You will receive the text content of a calibration certificate Excel file (converted from PDF).
The certificate may contain:
- 갑지 (Cover page): basic equipment info
- 을지 (Calibration results): measurement data (may be multiple pages)
- 적합성검토서 (Conformity Review): PASS/FAIL results (last page, optional)

Extract ALL available information and return a JSON object with these fields:
{
  "성적서번호": "certificate number",
  "고객명": "client/customer name",
  "장비명": "equipment description",
  "제조사": "manufacturer",
  "모델": "model name",
  "시리얼": "serial number",
  "관리번호": "identification/management number",
  "교정일": "date of calibration",
  "차기교정일": "next calibration date",
  "적합성검토": true/false (whether conformity review sheet exists),
  "측정포인트수": number of measurement points,
  "전체판정": "PASS" or "FAIL" or null,
  "측정요약": "brief summary of measurements in Korean"
}

Rules:
- Return ONLY the JSON object, no additional text
- If a field is not found, use null
- Dates should be in original format (e.g., "2026. 02. 05" or "2026-02-19")
- For 전체판정: PASS only if ALL measurement points passed, FAIL if any failed
- 측정요약: briefly describe what was measured and key results in Korean (1-2 sentences)
"""


def mistral_parse(excel_bytes):
    """Mistral API로 교정성적서 파싱"""
    text = excel_to_text(excel_bytes)

    # 토큰 절약: 텍스트가 너무 길면 잘라내기
    if len(text) > 8000:
        text = text[:8000] + '\n... (truncated)'

    prompt = f"""다음은 교정성적서 Excel 파일의 내용입니다. 정보를 추출해주세요.

{text}"""

    result = call_mistral(prompt, system_prompt=SYSTEM_PROMPT)

    # JSON 파싱
    try:
        parsed = json.loads(result['content'])
    except json.JSONDecodeError:
        # JSON 부분만 추출 시도
        content = result['content']
        start = content.find('{')
        end = content.rfind('}') + 1
        if start >= 0 and end > start:
            parsed = json.loads(content[start:end])
        else:
            raise Exception(f'JSON 파싱 실패: {content[:200]}')

    parsed['_llm_meta'] = {
        'elapsed': round(result['elapsed'], 2),
        'input_tokens': result['input_tokens'],
        'output_tokens': result['output_tokens'],
        'model': MISTRAL_MODEL,
    }

    return parsed


def compare_results(rule_based, llm_based, api_acpt):
    """규칙 기반 vs LLM 결과 비교"""
    fields = ['성적서번호', '고객명', '장비명', '제조사', '모델', '시리얼',
              '관리번호', '교정일', '차기교정일', '전체판정']

    print(f'\n  --- 비교: {api_acpt} ---')
    print(f'  {"필드":12s}  {"규칙기반":30s}  {"LLM":30s}  일치')
    print(f'  {"-"*90}')

    match_count = 0
    compare_count = 0

    for f in fields:
        rv = str(rule_based.get(f, '') or '').strip()
        lv = str(llm_based.get(f, '') or 'null').strip()

        # 정규화 비교
        rv_norm = rv.lower().replace(' ', '').replace('.', '').replace('-', '')
        lv_norm = lv.lower().replace(' ', '').replace('.', '').replace('-', '')

        if rv_norm == 'none' or rv_norm == '':
            rv_norm = ''
        if lv_norm == 'null' or lv_norm == 'none' or lv_norm == '':
            lv_norm = ''

        if rv_norm or lv_norm:
            compare_count += 1
            matched = rv_norm == lv_norm
            if matched:
                match_count += 1
            mark = 'O' if matched else 'X'
            # LLM이 규칙보다 더 많은 정보를 추출했으면 '+' 표시
            if not rv_norm and lv_norm:
                mark = '+'
        else:
            mark = '-'  # 둘 다 없음

        print(f'  {f:12s}  {rv[:30]:30s}  {lv[:30]:30s}  {mark}')

    rate = match_count / max(compare_count, 1) * 100
    print(f'\n  일치율: {match_count}/{compare_count} ({rate:.0f}%)')

    # LLM만 추출 성공한 필드 카운트
    llm_extra = 0
    for f in fields:
        rv = str(rule_based.get(f, '') or '').strip()
        lv = str(llm_based.get(f, '') or 'null').strip()
        if (not rv or rv == 'None') and lv and lv != 'null' and lv != 'None':
            llm_extra += 1

    if llm_extra > 0:
        print(f'  LLM 추가 추출: {llm_extra}개 필드')

    # 측정요약 출력
    if llm_based.get('측정요약'):
        print(f'  측정요약: {llm_based["측정요약"]}')

    return {
        'match_count': match_count,
        'compare_count': compare_count,
        'llm_extra': llm_extra,
    }


def run_test():
    """Mistral API 백업 파서 통합 테스트"""
    print('=' * 70)
    print('Mistral API 백업 파서 테스트')
    print('=' * 70)

    # 1. API 연결 테스트
    print('\n[1] Mistral API 연결 테스트...')
    try:
        result = call_mistral('Say "OK" in JSON: {"status": "OK"}')
        print(f'    연결 성공! 응답시간: {result["elapsed"]:.2f}초')
        print(f'    모델: {MISTRAL_MODEL}')
        print(f'    응답: {result["content"][:100]}')
    except Exception as e:
        print(f'    연결 실패: {e}')
        return

    # 2. 교정성적서 데이터 수집
    print('\n[2] 교정성적서 데이터 수집...')
    session = ktools_login(USER_ID, USER_PWD)
    session.get(f'{BASE_URL}/spm/contents/spm0907.do?cnsnClsIdx=32')
    all_data = fetch_all(session)

    completed = [d for d in all_data if '완료' in d.get('pgstNm', '')]
    print(f'    전체: {len(all_data)}건, 완료: {len(completed)}건')

    # 3. 다양한 케이스 샘플링
    #    - 적합성검토서 있는 건 + 없는 건 혼합
    #    - 다양한 장비 유형
    import random
    random.seed(123)  # 재현 가능

    # 먼저 10건 다운로드 + 규칙기반 파싱하여 적합성검토 있/없 분류
    print('\n[3] 샘플 10건 다운로드 & 분류...')
    candidates = random.sample(completed, min(30, len(completed)))

    has_conf = []     # 적합성검토서 있는 건
    no_conf = []      # 없는 건
    excels = {}       # api_acpt → excel_bytes 캐시

    for d in candidates:
        if len(has_conf) >= 5 and len(no_conf) >= 5:
            break

        api_acpt = make_api_acpt_no(d)
        excel = download_cert_excel(session, api_acpt)
        if not excel:
            continue

        cert = parse_cert_excel(excel)
        excels[api_acpt] = excel

        if cert['적합성검토'] and len(has_conf) < 5:
            has_conf.append((api_acpt, d, cert))
        elif not cert['적합성검토'] and len(no_conf) < 5:
            no_conf.append((api_acpt, d, cert))

        time.sleep(0.3)

    samples = has_conf + no_conf
    print(f'    적합성검토 있음: {len(has_conf)}건, 없음: {len(no_conf)}건')
    print(f'    총 테스트: {len(samples)}건')

    # 4. Mistral API 파싱 테스트
    print(f'\n[4] Mistral API 파싱 테스트 (vs 규칙기반)')
    print(f'    모델: {MISTRAL_MODEL}')
    print()

    total_stats = {
        'match_total': 0,
        'compare_total': 0,
        'llm_extra_total': 0,
        'llm_times': [],
        'llm_tokens': [],
    }

    for i, (api_acpt, d, rule_cert) in enumerate(samples):
        prd = d.get('entpPrdNm', '')[:35]
        conf_mark = 'O' if rule_cert['적합성검토'] else 'X'
        print(f'  [{i+1}/{len(samples)}] {api_acpt:22s}  적합{conf_mark}  {prd}')

        try:
            llm_cert = mistral_parse(excels[api_acpt])
            meta = llm_cert.get('_llm_meta', {})
            print(f'    LLM: {meta.get("elapsed", 0):.1f}초, '
                  f'입력={meta.get("input_tokens", 0)}, 출력={meta.get("output_tokens", 0)} 토큰')

            total_stats['llm_times'].append(meta.get('elapsed', 0))
            total_stats['llm_tokens'].append(
                meta.get('input_tokens', 0) + meta.get('output_tokens', 0))

            comp = compare_results(rule_cert, llm_cert, api_acpt)
            total_stats['match_total'] += comp['match_count']
            total_stats['compare_total'] += comp['compare_count']
            total_stats['llm_extra_total'] += comp['llm_extra']

        except Exception as e:
            print(f'    LLM 에러: {e}')

        time.sleep(1)  # API rate limit 배려
        print()

    # 5. 종합 결과
    print(f'\n{"="*70}')
    print(f'종합 결과')
    print(f'{"="*70}')

    total_match = total_stats['match_total']
    total_compare = total_stats['compare_total']
    rate = total_match / max(total_compare, 1) * 100

    print(f'\n--- 일치율 ---')
    print(f'  규칙기반 vs LLM 일치: {total_match}/{total_compare} ({rate:.1f}%)')
    print(f'  LLM 추가 추출:       {total_stats["llm_extra_total"]}개 필드')

    if total_stats['llm_times']:
        times = total_stats['llm_times']
        tokens = total_stats['llm_tokens']
        print(f'\n--- LLM 성능 ---')
        print(f'  평균 응답시간: {sum(times)/len(times):.2f}초')
        print(f'  최소/최대:     {min(times):.2f}초 / {max(times):.2f}초')
        print(f'  평균 토큰:     {sum(tokens)//len(tokens)}')
        print(f'  총 토큰:       {sum(tokens)}')

    # 비용 추정 (mistral-small: ~$0.1/M input, $0.3/M output — 추정치)
    total_tokens = sum(total_stats['llm_tokens'])
    est_cost_per_1k = total_tokens / max(len(total_stats['llm_tokens']), 1) * 0.0002
    print(f'\n--- 비용 추정 (1,000건 기준) ---')
    print(f'  예상 토큰:     {total_tokens / max(len(total_stats["llm_tokens"]), 1) * 1000:.0f}')
    print(f'  예상 비용:     ~${est_cost_per_1k * 1000:.2f}')

    print(f'\n{"="*70}')
    print(f'결론: LLM 백업 파서 {"적합" if rate >= 70 else "추가 개선 필요"}')
    print(f'{"="*70}')


if __name__ == '__main__':
    run_test()
