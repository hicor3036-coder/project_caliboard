# -*- coding: utf-8 -*-
"""
Created on Sat Feb 22 2026

@author: hicor
교정성적서 Excel 다운로드 & 파싱

=== 전체 흐름 ===
  k-tools 로그인(세션)
    → 보안 토큰 발급 (get_token)
    → PDF→Excel 변환 요청 (서버에서 변환)
    → Excel 다운로드 (download_cert_excel)
    → 규칙기반 파싱 (parse_cert_excel)
    → [필요 시] LLM 백업 파싱 (_llm_supplement)

=== 교정성적서 구조 (PDF→Excel 변환 결과) ===
  - 갑지 (Page 1):  표지. 장비 기본 정보 (성적서번호, 고객명, 장비명, 제조사/모델, 시리얼 등)
  - 을지 (Page 2~): 교정 측정 결과. 비정형 데이터라 규칙기반 파싱 불가 → LLM에 위임
  - 적합성검토서 (마지막 페이지, 선택): Conformity Review. 장비정보 + PASS/FAIL 측정 결과
    ※ 적합성검토서가 없는 성적서도 많음 (약 50%)

=== 3단계 fallback 파싱 전략 ===
  1차: 규칙기반 (로컬 로직) — 갑지 + 적합성검토서에서 정형 필드 추출
  2차: Mistral API (mistral-small) — 핵심 필드 2개 이상 누락 시 발동, 최대 2회 재시도
  3차: Groq API (llama-3.3-70b) — Mistral 실패(429 등) 시 최후의 보루, 최대 1회 재시도

=== 유의사항 ===
  - acptNo 변환: DB는 zero-padded (예: 02-012), API는 unpadded (예: 2-12) → make_api_acpt_no()
  - 갑지 파싱 시 영문 라벨("Serial Number" 등)이 값으로 잘못 추출되는 케이스 있음 → BAD_VALUES로 후처리
  - 적합성검토서 헤더는 2가지 레이아웃(패턴A/B)이 있음 → _parse_conformity() 내 주석 참고
  - LLM은 빈 필드만 채움 (규칙기반 결과 우선). 기존 값은 절대 덮어쓰지 않음
  - LLM 호출 시 Excel 전체를 텍스트로 변환 → 8,000자 초과 시 truncate
  - Mistral 무료 티어: rate limit 빡빡 (1 req/sec, 30 req/min) → 429 시 exponential backoff
  - Groq 무료 티어: 일일 토큰 무제한이지만 요청 수 제한 (250 req/day, 30 req/min)
  - API 키는 현재 이 파일에 하드코딩됨. 추후 환경변수나 별도 config로 분리 권장

=== 외부 모듈 의존성 ===
  - ktools_login.py: k-tools 로그인 세션 생성
  - ktools_수집.py:  장비 목록 전체 수집 (fetch_all)
  - openpyxl:        Excel 파일 읽기
  - requests:        HTTP API 호출 (k-tools, Mistral, Groq)

=== 주요 함수 (외부 호출용) ===
  - make_api_acpt_no(item) → API용 접수번호 문자열
  - download_cert_excel(session, api_acpt_no) → Excel 바이트 또는 None
  - parse_cert_excel(excel_bytes, use_llm=True) → 구조화된 dict
"""

import sys
import io
import json
import time
import requests
import openpyxl
from io import BytesIO
from ktools_login import ktools_login

# Windows 콘솔 인코딩 문제 방지
# - Windows cmd/PowerShell은 cp949라 한글 깨짐 → UTF-8로 강제
# - Spyder IDE의 TTYOutStream은 buffer 속성이 없어서 hasattr 체크 필요
if sys.platform == 'win32' and hasattr(sys.stdout, 'buffer') and not isinstance(sys.stdout, io.TextIOWrapper):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

USER_ID = 'hicor'
USER_PWD = 'dlacodnr1!'
BASE_URL = 'https://k-tools.ktl.re.kr'

# =============================================================================
# LLM API 설정 (TODO: 추후 환경변수 또는 config 파일로 분리)
# =============================================================================

# 2차 fallback: Mistral
# - 무료 토큰 넉넉하지만 rate limit 빡빡 (1 req/sec, 30 req/min)
# - 429 발생 시 exponential backoff로 최대 2회 재시도
# - OpenAI 호환 API 형식 (chat/completions)
MISTRAL_API_KEY = '8GmClOHJYZr3PpZUAunyzQykOJPghT8D'
MISTRAL_URL = 'https://api.mistral.ai/v1/chat/completions'
MISTRAL_MODEL = 'mistral-small-latest'

# 3차 fallback: Groq (Mistral 429 실패 시 최후의 보루)
# - Groq LPU 칩 기반이라 응답 매우 빠름 (~1초)
# - 무료 티어: 일일 토큰 무제한, 요청 수 제한 (250 req/day)
# - OpenAI 호환 API 형식 (chat/completions)
GROQ_API_KEY = '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'
GROQ_URL = 'https://api.groq.com/openai/v1/chat/completions'
GROQ_MODEL = 'llama-3.3-70b-versatile'

# LLM fallback 발동 기준
# - 아래 핵심 필드 중 THRESHOLD 이상 누락 시 LLM 호출
# - 적합성검토서가 있는 성적서는 규칙기반으로 대부분 추출 → LLM 불필요
# - 적합성검토서가 없는 성적서는 갑지만으로 부족한 경우가 많아 LLM 발동률 높음
LLM_MISSING_THRESHOLD = 2
LLM_KEY_FIELDS = ['제조사', '모델', '시리얼', '교정일']


# =============================================================================
# API 함수
# =============================================================================

def get_token(session):
    """보안 토큰 발급"""
    res = session.post(f'{BASE_URL}/spm/api/getSecToken')
    data = res.json()
    if data.get('code') != 200:
        raise Exception(f'토큰 발급 실패: {data}')
    return data['data']['token']


def make_api_acpt_no(item):
    """DB 데이터에서 API용 접수번호 생성 (끝자리 zero-padding 제거)
    DB: acptNo='26-010119-02-012' → API: '26-010119-02-12'
    """
    return f'{item["incsRcpnSrno"]}-{int(item["rcpnArtcSrno"])}'


def download_cert_excel(session, api_acpt_no):
    """교정성적서 Excel 다운로드
    1) PDF→Excel 변환 요청 (서버에서 MarkAny DRM 해제 + PDF→Excel 변환)
    2) 변환된 Excel 파일 다운로드
    반환: Excel 바이트 또는 None (실패 시)

    ※ 실패 케이스:
      - DRM 해제 불가 (서버에서 code != 200)
      - MarkAny 라이선스 만료
      - 성적서 원본 PDF가 존재하지 않는 경우
    """
    token = get_token(session)

    # Step 1: 변환 요청 — 서버에서 DRM 해제 + PDF→Excel 변환
    # ※ 변환은 세션 단위. 이전 변환 결과가 남아있으면 덮어씀
    convert_res = session.post(
        f'{BASE_URL}/spm/api/spm0907_saveReportCardPdfToExcel.ajax',
        data=f'acptNo={api_acpt_no}&token={token}',
    )
    result = convert_res.json()
    if result.get('code') != 200:
        return None

    # Step 2: 다운로드 — Step 1에서 변환된 Excel을 가져옴
    # ※ Content-Type이 spreadsheet가 아니면 HTML 에러 페이지일 수 있음
    download_res = session.get(f'{BASE_URL}/excel/getAcptNoPdfToExcel.do')
    content_type = download_res.headers.get('Content-Type', '')
    if 'spreadsheet' not in content_type:
        return None

    return download_res.content


# =============================================================================
# LLM 백업 파서 (3단계 fallback: Mistral → Groq)
# =============================================================================

_LLM_SYSTEM_PROMPT = """You are a calibration certificate data extraction assistant.
You will receive the text content of a calibration certificate Excel file (converted from PDF).
The certificate may contain:
- 갑지 (Cover page): basic equipment info
- 을지 (Calibration results): measurement data (may be multiple pages)
- 적합성검토서 (Conformity Review): PASS/FAIL results (last page, optional)

Extract ALL available information and return a JSON object with these fields:
{
  "성적서번호": "certificate number",
  "고객명": "client/customer name",
  "장비명": "equipment description",
  "제조사": "manufacturer",
  "모델": "model name",
  "시리얼": "serial number",
  "관리번호": "identification/management number",
  "교정일": "date of calibration",
  "차기교정일": "next calibration date",
  "전체판정": "PASS" or "FAIL" or null,
  "측정요약": "brief summary of measurements in Korean (1-2 sentences)"
}

Rules:
- Return ONLY the JSON object, no additional text
- If a field is not found, use null
- Dates should be in original format
- For 전체판정: PASS only if ALL measurement points passed, FAIL if any failed"""

# LLM 프로바이더 설정 (순서대로 시도, 실패 시 다음으로 fallback)
# - retries: 429 rate limit 시 재시도 횟수 (exponential backoff: 1초→3초→5초)
# - 새 프로바이더 추가 시 이 리스트에 dict 추가하면 됨 (OpenAI 호환 API만 가능)
_LLM_PROVIDERS = [
    {
        'name': 'Mistral',
        'url': MISTRAL_URL,
        'key': MISTRAL_API_KEY,
        'model': MISTRAL_MODEL,
        'retries': 2,
    },
    {
        'name': 'Groq',
        'url': GROQ_URL,
        'key': GROQ_API_KEY,
        'model': GROQ_MODEL,
        'retries': 1,
    },
]


def _excel_to_text(excel_bytes):
    """Excel 바이트 → 시트별 텍스트 변환 (LLM 입력용)
    각 시트를 '=== Sheet: 이름 ===' 구분자로 나누고, 셀을 ' | '로 연결
    ※ 빈 행은 스킵하여 토큰 절약
    """
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    parts = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        lines = [f'=== Sheet: {sn} ===']
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if not any(vals):
                continue
            lines.append(' | '.join(vals))
        parts.append('\n'.join(lines))
    wb.close()
    return '\n\n'.join(parts)


def _call_llm(provider, prompt, system_prompt=None):
    """단일 LLM 프로바이더 호출 (429 재시도 포함)
    반환: 응답 텍스트 또는 Exception 발생

    ※ OpenAI 호환 API 형식 — Mistral, Groq 모두 동일 포맷
    ※ response_format: json_object → 응답이 반드시 JSON으로 옴
    ※ temperature=0.0 → 동일 입력에 대해 일관된 결과
    ※ 429 재시도: wait = 2^attempt + 1 (1초 → 3초 → 5초)
    """
    messages = []
    if system_prompt:
        messages.append({'role': 'system', 'content': system_prompt})
    messages.append({'role': 'user', 'content': prompt})

    payload = {
        'model': provider['model'],
        'messages': messages,
        'temperature': 0.0,
        'max_tokens': 2000,
        'response_format': {'type': 'json_object'},
    }
    headers = {
        'Authorization': f'Bearer {provider["key"]}',
        'Content-Type': 'application/json',
    }

    retries = provider.get('retries', 2)
    for attempt in range(retries + 1):
        res = requests.post(provider['url'], json=payload, headers=headers, timeout=30)
        if res.status_code == 429:
            wait = 2 ** attempt + 1
            time.sleep(wait)
            continue
        if res.status_code != 200:
            raise Exception(f'{provider["name"]} {res.status_code}: {res.text[:200]}')
        data = res.json()
        return data['choices'][0]['message']['content']

    raise Exception(f'{provider["name"]} 429: rate limit {retries+1}회 초과')


def _llm_parse(excel_bytes):
    """LLM으로 교정성적서 파싱 (3단계 fallback)

    2차: Mistral (최대 2회 재시도)
    3차: Groq Llama 3.3 70B (최후의 보루)

    반환: (dict, provider_name) 또는 (None, None)

    ※ 텍스트 8,000자 초과 시 truncate — 을지(측정 데이터)가 길면 잘릴 수 있으나
      갑지 + 적합성검토서 정보는 앞쪽에 위치하므로 대부분 포함됨
    ※ JSON 파싱 실패 시 응답에서 { } 범위를 찾아 재시도
    """
    text = _excel_to_text(excel_bytes)
    if len(text) > 8000:
        text = text[:8000] + '\n... (truncated)'

    prompt = f'다음은 교정성적서 Excel 파일의 내용입니다. 정보를 추출해주세요.\n\n{text}'

    for provider in _LLM_PROVIDERS:
        try:
            content = _call_llm(provider, prompt, system_prompt=_LLM_SYSTEM_PROMPT)

            # JSON 파싱
            try:
                return json.loads(content), provider['name']
            except json.JSONDecodeError:
                start = content.find('{')
                end = content.rfind('}') + 1
                if start >= 0 and end > start:
                    return json.loads(content[start:end]), provider['name']
        except Exception:
            continue  # 다음 프로바이더로 fallback

    return None, None


def _llm_supplement(result, excel_bytes):
    """규칙기반 결과에 LLM으로 누락 필드 보강

    핵심 필드 중 LLM_MISSING_THRESHOLD 이상 누락 시 LLM 호출.
    LLM 결과로 빈 필드만 채움 (기존 규칙기반 결과 우선 — 절대 덮어쓰지 않음).
    반환: (보강된 result, llm_used: bool)

    ※ result에 추가되는 메타 필드:
      - _llm_보강:     LLM이 채운 필드명 리스트 (예: ['제조사', '모델'])
      - _llm_provider: 사용된 프로바이더명 (예: 'Mistral' 또는 'Groq')
      - 측정요약:      LLM이 생성한 1~2문장 한글 요약 (규칙기반에는 없는 필드)
    """
    missing = [f for f in LLM_KEY_FIELDS if not result.get(f)]
    if len(missing) < LLM_MISSING_THRESHOLD:
        return result, False

    llm, provider = _llm_parse(excel_bytes)
    if not llm:
        return result, False

    # 누락 필드만 LLM으로 보강
    FILL_FIELDS = ['성적서번호', '고객명', '장비명', '제조사', '모델',
                   '시리얼', '관리번호', '교정일', '차기교정일', '전체판정']
    filled = []
    for f in FILL_FIELDS:
        if not result.get(f) and llm.get(f) and str(llm[f]) != 'null':
            result[f] = str(llm[f])
            filled.append(f)

    # 측정요약은 항상 LLM에서 가져옴 (규칙기반에는 없는 필드)
    if llm.get('측정요약') and str(llm['측정요약']) != 'null':
        result['측정요약'] = str(llm['측정요약'])

    result['_llm_보강'] = filled
    result['_llm_provider'] = provider
    return result, True


# =============================================================================
# 규칙기반 파싱 함수
# ※ Excel 시트 구조는 PDF→Excel 변환기에 의해 결정되므로 셀 위치가 유동적
# ※ 라벨(영문) 기반으로 값을 찾는 방식 → 라벨 위치가 바뀌면 파싱 실패 가능
# ※ 파싱 실패 시 LLM fallback이 보완하므로 100% 정확할 필요는 없음
# =============================================================================

def _find_conformity_sheet(wb):
    """적합성검토서 시트 찾기 (마지막 시트에서 역순 탐색)
    ※ 적합성검토서는 보통 마지막 시트에 있으므로 역순으로 찾음
    ※ 첫 3행 내에 'CONFORMITY' 문자열 포함 여부로 판단
    """
    for sn in reversed(wb.sheetnames):
        ws = wb[sn]
        for row in ws.iter_rows(max_row=3, values_only=True):
            joined = ' '.join(str(v) for v in row if v).upper()
            if 'CONFORMITY' in joined:
                return ws
    return None


def _parse_cover(wb):
    """갑지 (Page 1) 파싱 — 기본 정보 추출

    ※ 갑지 레이아웃 예시:
      Certificate No.: 25-074958-01-265
      Client
        Name:                    KOREA AEROSPACE INDUSTRIES, LTD.
      Description:               ACCELEROMETER
      Manufacturer and Model:    PCB / 352C33
      Serial Number:             SN-LW331510 [Identification Number: K30917]
      Date of Calibration:       13   November   2025

    ※ 셀 위치가 고정이 아님 — 라벨 오른쪽(c+1, c+2, c+3 등)에서 값을 찾음
    ※ 한글 양식도 있어서 '성적서 번호', '교정일자' 등 한글 라벨도 처리
    """
    info = {}
    ws = wb['Page 1'] if 'Page 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    # 모든 셀을 (행, 열) → 값 딕셔너리로 변환
    # ※ values_only=False로 셀 좌표를 유지해야 인접 셀 탐색 가능
    cells = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None:
                cells[(cell.row, cell.column)] = str(cell.value).strip()

    found_client = False
    for (r, c), val in sorted(cells.items()):
        # 성적서번호
        if 'Certificate No' in val and ':' in val:
            info['성적서번호'] = val.split(':')[-1].strip()
        elif '성적서 번호' in val and ':' in val:
            info['성적서번호'] = val.split(':')[-1].strip()

        # Client 섹션 감지
        if 'Client' in val:
            found_client = True

        # 고객명 — Client 섹션 아래의 Name
        if found_client and ('Name' in val and 'Model' not in val and ':' in val):
            name_val = cells.get((r, c + 2)) or cells.get((r, c + 1))
            if name_val and '고객명' not in info:
                info['고객명'] = name_val
            found_client = False  # 첫 번째 Name만

        # 장비명
        if 'Description' in val and ':' in val:
            desc_val = cells.get((r, c + 2)) or cells.get((r, c + 1))
            if desc_val and '장비명' not in info:
                info['장비명'] = desc_val

        # 제조사/모델
        if 'Manufacturer and Model' in val:
            model_val = cells.get((r, c + 4)) or cells.get((r, c + 3)) or cells.get((r, c + 2))
            if model_val:
                parts = model_val.split('/')
                info['제조사'] = parts[0].strip()
                info['모델'] = parts[1].strip() if len(parts) > 1 else ''

        # 시리얼
        if val.startswith('Serial Number') or val.startswith('시리얼') or '일련번호' in val:
            sn_val = cells.get((r, c + 2)) or cells.get((r, c + 1))
            if sn_val and '시리얼' not in info:
                info['시리얼'] = sn_val.split('[')[0].strip()
                if '[' in str(sn_val):
                    info['관리번호'] = str(sn_val).split(':')[-1].strip().rstrip(']')

        # 교정일
        if 'Date of Calibration' in val or '교정일자' in val:
            date_val = cells.get((r, c + 3)) or cells.get((r, c + 2)) or cells.get((r, c + 1))
            if date_val and '교정일' not in info:
                info['교정일'] = str(date_val).strip()

    # 후처리: 영문 라벨이 값으로 잘못 파싱된 경우 제거
    # ※ 한글 양식 성적서에서 발생 — 영문 라벨이 별도 셀에 있어 값 위치로 잘못 인식
    # ※ 예: 제조사='Serial Number', 시리얼='The due date' 등
    BAD_VALUES = {'Serial Number', 'The due date', 'Manufacturer', 'Model',
                  'Description', 'Date of Calibration', 'Certificate No',
                  'Identification Number', 'Client', 'Name'}
    for key in list(info.keys()):
        if info[key] in BAD_VALUES:
            del info[key]

    return info


def _parse_conformity(ws):
    """적합성검토서 파싱 — 장비정보 + 측정결과 + PASS/FAIL

    ※ 적합성검토서는 2개 영역으로 구성:
      1) 장비 헤더: Manufacturer, Model, Description, Serial Number 등
      2) 측정 데이터: 각 포인트별 측정값 + PASS/FAIL 판정

    ※ 헤더 레이아웃이 2가지 패턴으로 나뉨 (아래 주석 참고)
    ※ 측정 데이터는 PASS/FAIL 문자열이 포함된 행만 추출
    """
    info = {}
    measurements = []

    rows = list(ws.iter_rows(values_only=True))

    # --- 장비 헤더 영역 파싱 ---
    # 적합성검토서 헤더는 2가지 레이아웃이 있음:
    # 패턴A: ['DYTRAN', 'Model', '3273A2', 'Description', 'Vibration transducers']
    #         ['Manufacturer']
    # 패턴B: ['Manufacturer', 'KULITE', 'Model', 'HKL-375-100A', 'Description', '...']
    #
    # 공통: Manufacturer/Model/Description 라벨과 값이 같은 행 또는 인접 행에 있음

    HEADER_LABELS = {'Manufacturer', 'Model', 'Description', 'Serial Number',
                     'Certificate No.', 'Identification', 'Number'}

    for idx, row in enumerate(rows):
        vals = [str(v).strip() if v is not None else '' for v in row]
        joined = ' '.join(vals)

        # Manufacturer/Model/Description 처리
        if 'Manufacturer' in vals or 'Model' in vals:
            # 이 행 + 바로 위 행을 합쳐서 분석 (패턴A 대응)
            combined = list(vals)
            if idx > 0:
                prev = [str(v).strip() if v is not None else '' for v in rows[idx - 1]]
                # 위 행에 라벨 없이 값만 있으면 = 패턴A
                if not any(v in HEADER_LABELS for v in prev):
                    combined = prev + combined

            # 라벨-값 쌍 추출
            for i, v in enumerate(combined):
                if v == 'Manufacturer' and '제조사' not in info:
                    # 두 가지 레이아웃:
                    # 패턴A (Manufacturer 단독 행):
                    #   행1: [제조사, 'Model', 모델, 'Description', 장비명]
                    #   행2: ['Manufacturer']
                    #   → 위 행에서 찾기
                    # 패턴B (같은 행):
                    #   행: ['Manufacturer', 제조사, 'Model', 모델, 'Description']
                    #   → 같은 행에서 Manufacturer 바로 뒤 값

                    # 같은 행(vals)에 Manufacturer 외에 Model도 있으면 → 패턴B
                    if 'Model' in vals:
                        # 패턴B: 같은 행에서 Manufacturer와 Model 사이의 값
                        mfr_idx = vals.index('Manufacturer')
                        model_idx = vals.index('Model')
                        for j in range(mfr_idx + 1, model_idx):
                            if vals[j] and vals[j] not in HEADER_LABELS:
                                info['제조사'] = vals[j]
                                break
                    else:
                        # 패턴A: 위 행에서 찾기
                        for j in range(i - 1, -1, -1):
                            if combined[j] and combined[j] not in HEADER_LABELS:
                                info['제조사'] = combined[j]
                                break

                if v == 'Model' and '모델' not in info:
                    # 같은 행(vals)에서 Model 뒤의 첫 비-라벨 값
                    if v in vals:
                        m_idx = vals.index('Model')
                        for j in range(m_idx + 1, len(vals)):
                            if vals[j] and vals[j] not in HEADER_LABELS:
                                info['모델'] = vals[j]
                                break

                if v == 'Description' and '장비명' not in info:
                    # 같은 행(vals)에서 Description 뒤의 첫 비-라벨 값
                    if v in vals:
                        d_idx = vals.index('Description')
                        for j in range(d_idx + 1, len(vals)):
                            if vals[j] and vals[j] not in HEADER_LABELS:
                                info['장비명'] = vals[j]
                                break

        if 'Serial Number' in joined:
            for i, v in enumerate(vals):
                if 'Serial Number' in v or v == 'Serial Number':
                    for j in range(i + 1, len(vals)):
                        if vals[j] and vals[j] != 'Serial Number':
                            info['시리얼'] = vals[j]
                            break
                    break

        if 'Certificate No' in joined:
            for i, v in enumerate(vals):
                if 'Certificate No' in v:
                    if ':' in v:
                        info['성적서번호'] = v.split(':')[-1].strip()
                    else:
                        for j in range(i + 1, len(vals)):
                            if vals[j]:
                                info['성적서번호'] = vals[j]
                                break
                    break

        if 'Identification' in joined:
            for i, v in enumerate(vals):
                if 'Identification' in v:
                    for j in range(i + 1, len(vals)):
                        if vals[j] and vals[j] not in ('Number', 'Identification Number'):
                            info['관리번호'] = vals[j]
                            break
                    break

        if 'Date of Calibration' in joined or 'Calibration' in joined:
            for i, v in enumerate(vals):
                if 'Calibration' in v and 'Date' in joined:
                    for j in range(i + 1, len(vals)):
                        if vals[j] and any(c.isdigit() for c in vals[j]):
                            info['교정일'] = vals[j]
                            # 차기교정일은 그 다음 날짜
                            for k in range(j + 1, len(vals)):
                                if vals[k] and any(c.isdigit() for c in vals[k]):
                                    info['차기교정일'] = vals[k]
                                    break
                            break
                    break

    # --- 측정 데이터 영역 파싱 ---
    # PASS/FAIL이 있는 행을 데이터 행으로 인식
    for row in rows:
        vals = [v for v in row]
        str_vals = [str(v).strip() if v is not None else '' for v in vals]

        # PASS 또는 FAIL이 포함된 행 = 측정 데이터
        if 'PASS' not in str_vals and 'FAIL' not in str_vals:
            continue

        # 숫자 데이터 추출
        numbers = []
        conformity = ''
        non_empty = []
        for v in vals:
            sv = str(v).strip() if v is not None else ''
            if sv in ('PASS', 'FAIL'):
                conformity = sv
            elif sv and sv != '-' and sv != 'None':
                non_empty.append(sv)
                try:
                    numbers.append(float(sv.replace(' ', '').replace(',', '')))
                except ValueError:
                    pass

        if conformity:
            measurements.append({
                '원본데이터': non_empty,
                '숫자값': numbers,
                '판정': conformity,
            })

    info['측정결과'] = measurements
    info['측정포인트수'] = len(measurements)
    if measurements:
        pass_cnt = sum(1 for m in measurements if m['판정'] == 'PASS')
        fail_cnt = sum(1 for m in measurements if m['판정'] == 'FAIL')
        info['PASS'] = pass_cnt
        info['FAIL'] = fail_cnt
        info['전체판정'] = 'PASS' if fail_cnt == 0 else 'FAIL'
    else:
        info['전체판정'] = None

    return info


def _cross_validate(cover, conf):
    """갑지 vs 적합성검토서 교차검증 — 불일치 항목 검출

    검증 대상: 성적서번호, 제조사, 모델, 시리얼, 관리번호
    반환: 불일치 목록 [{'항목': ..., '갑지': ..., '적합성검토': ...}, ...]
    """
    CHECK_FIELDS = [
        ('성적서번호', '성적서번호'),
        ('제조사', '제조사'),
        ('모델', '모델'),
        ('시리얼', '시리얼'),
        ('관리번호', '관리번호'),
    ]

    mismatches = []
    for field, label in CHECK_FIELDS:
        cover_val = cover.get(field, '').strip()
        conf_val = conf.get(field, '').strip()

        # 둘 다 없으면 비교 불가 → skip
        if not cover_val or not conf_val:
            continue

        # 정규화 비교 (대소문자, 공백 무시)
        if cover_val.lower().replace(' ', '') != conf_val.lower().replace(' ', ''):
            mismatches.append({
                '항목': label,
                '갑지': cover_val,
                '적합성검토': conf_val,
            })

    return mismatches


def parse_cert_excel(excel_bytes, use_llm=True):
    """교정성적서 Excel 파싱 → 구조화된 dict 반환 (메인 진입점)

    Args:
        excel_bytes: download_cert_excel()로 받은 Excel 바이트
        use_llm:     True면 핵심 필드 누락 시 LLM fallback 발동 (기본값 True)
                     대량 처리 시 False로 끄면 API 비용/시간 절약

    Returns:
        dict — 주요 키:
          성적서번호, 고객명, 장비명, 제조사, 모델, 시리얼, 관리번호,
          교정일, 차기교정일, 적합성검토(bool), 전체판정(PASS/FAIL/None),
          측정포인트수, 측정결과(list), 불일치(list),
          _llm_보강(list), _llm_provider(str), 측정요약(str)

    파싱 전략:
      1차: 규칙기반 — 갑지 + 적합성검토서
      2차: Mistral LLM — 핵심 필드 2개 이상 누락 시
      3차: Groq Llama — Mistral 실패 시 최후의 보루
    """
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))

    result = {
        '시트수': len(wb.sheetnames),
        '시트목록': wb.sheetnames,
    }

    # 갑지 파싱
    cover = _parse_cover(wb)
    result.update(cover)

    # 적합성검토서 파싱
    conformity_ws = _find_conformity_sheet(wb)
    if conformity_ws:
        result['적합성검토'] = True
        conf = _parse_conformity(conformity_ws)

        # 교차검증
        mismatches = _cross_validate(cover, conf)
        result['불일치'] = mismatches

        # 최종 값 결정: 갑지 우선, 적합성검토서로 보강
        for key in ('장비명', '교정일', '차기교정일', '성적서번호'):
            if key in conf and conf[key] and key not in result:
                result[key] = conf[key]
        if '차기교정일' in conf and conf['차기교정일']:
            result['차기교정일'] = conf['차기교정일']
        if '관리번호' in conf and conf['관리번호'] and '관리번호' not in cover:
            result['관리번호'] = conf['관리번호']

        result['측정결과'] = conf['측정결과']
        result['측정포인트수'] = conf['측정포인트수']
        result['PASS'] = conf.get('PASS', 0)
        result['FAIL'] = conf.get('FAIL', 0)
        result['전체판정'] = conf['전체판정']
    else:
        result['적합성검토'] = False
        result['불일치'] = []
        result['측정결과'] = []
        result['측정포인트수'] = 0
        result['전체판정'] = None

    wb.close()

    # LLM 백업 파싱 (핵심 필드 누락 시)
    # ※ _llm_보강이 빈 리스트면 규칙기반만으로 충분했다는 의미
    # ※ _llm_provider가 'Mistral' 또는 'Groq'면 해당 LLM이 보강한 것
    result['_llm_보강'] = []
    if use_llm:
        result, _ = _llm_supplement(result, excel_bytes)

    return result


# =============================================================================
# 테스트 실행 (직접 실행 시)
# - Spyder에서 그대로 실행 가능
# - 랜덤 10건을 뽑아 규칙기반 + LLM fallback 통합 테스트
# - 각 건마다 [Mistral+5] 또는 [Groq+3] 형태로 어떤 LLM이 몇 개 필드를 보강했는지 표시
# =============================================================================

if __name__ == '__main__':
    import random
    from ktools_수집 import fetch_all

    print('=' * 70)
    print('교정성적서 Excel 다운로드 & 파싱 테스트 (LLM 백업 통합)')
    print('=' * 70)

    # 1. 로그인 & 데이터 수집
    # ※ session.get(spm0907.do)는 교정 관리 메뉴 접근용 — 이걸 해야 API 호출 가능
    print('\n[1] 로그인 & 데이터 수집...')
    session = ktools_login(USER_ID, USER_PWD)
    session.get(f'{BASE_URL}/spm/contents/spm0907.do?cnsnClsIdx=32')
    all_data = fetch_all(session)

    completed = [d for d in all_data if '완료' in d.get('pgstNm', '')]
    print(f'    전체: {len(all_data)}건, 완료(성적서有): {len(completed)}건')

    # 2. 샘플링: 다양한 장비 유형 10건
    random.seed(42)
    samples = random.sample(completed, min(10, len(completed)))
    print(f'\n[2] 랜덤 {len(samples)}건 테스트 (LLM fallback 포함)\n')

    success = 0
    fail = 0
    llm_count = 0

    for i, d in enumerate(samples):
        api_acpt = make_api_acpt_no(d)
        prd = d.get('entpPrdNm', '')[:40]

        excel = download_cert_excel(session, api_acpt)
        if not excel:
            print(f'  [{i+1}] NG  {api_acpt:22s}  다운로드 실패  {prd}')
            fail += 1
            time.sleep(0.3)
            continue

        cert = parse_cert_excel(excel, use_llm=True)
        success += 1

        llm_fields = cert.get('_llm_보강', [])
        if llm_fields:
            llm_count += 1

        # 결과 출력
        conf_mark = 'O' if cert['적합성검토'] else 'X'
        provider = cert.get('_llm_provider', '')
        llm_mark = f' [{provider}+{len(llm_fields)}]' if llm_fields else ''
        print(f'  [{i+1}] {api_acpt:22s}  적합{conf_mark}{llm_mark}  {prd}')
        print(f'       성적서번호={cert.get("성적서번호", "?")}')
        print(f'       제조사={cert.get("제조사", "?")} / 모델={cert.get("모델", "?")}')
        print(f'       시리얼={cert.get("시리얼", "?")} / 관리번호={cert.get("관리번호", "?")}')
        print(f'       교정일={cert.get("교정일", "?")} / 차기={cert.get("차기교정일", "?")}')
        print(f'       판정={cert.get("전체판정", "-")} / 측정={cert.get("측정포인트수", 0)}pt')
        if llm_fields:
            print(f'       {provider} 보강: {", ".join(llm_fields)}')
        if cert.get('측정요약'):
            print(f'       측정요약: {cert["측정요약"]}')
        print()
        time.sleep(0.3)

    print(f'{"="*70}')
    print(f'결과: 성공={success}, 실패={fail}, LLM 보강={llm_count}건')
    print(f'{"="*70}')
